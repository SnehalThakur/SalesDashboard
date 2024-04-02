"""
Copyright (c) 2024 - Bizware International
"""

import logging
import os
import io
import tempfile
from pathlib import Path

import botocore
from openpyxl import Workbook, load_workbook
from io import BytesIO
import pandas as pd
import boto3
from botocore.exceptions import ClientError
from loguru import logger


def getS3Resource():
    # Let's use Amazon S3
    s3 = boto3.resource('s3', aws_access_key_id='AKIA3Z4VJEZ3KMT2DTGS',
                        aws_secret_access_key='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')
    return s3


def getS3Client():
    # Let's use Amazon S3
    s3 = boto3.client('s3', aws_access_key_id='AKIA3Z4VJEZ3KMT2DTGS',
                      aws_secret_access_key='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')
    return s3


def printBucketNames():
    s3 = getS3Resource()
    # Print out bucket names
    for bucket in s3.buckets.all():
        print(bucket.name)


# session = boto3.Session(
#     aws_access_key_id='ACCESS_KEY',
#     aws_secret_access_key='SECRET_KEY',
# )

KB = 1024
MB = 1024 * KB

SUPPORTED_FILE_TYPES = {
    'image/png': 'png',
    'image/jpeg': 'jpg',
    'application/pdf': 'pdf'
}


def getBucket(AWS_BUCKET='bizwarebucket'):
    s3 = getS3Resource()
    bucket = s3.Bucket(AWS_BUCKET)
    print("bucket -", bucket)
    return bucket


async def s3_upload(contents: bytes, key: str):
    logger.info(f'Uploading {key} to s3')
    bucket = getBucket()
    bucket.put_object(Key=key, Body=contents)


async def s3_download(key: str):
    try:
        s3 = getS3Resource()
        return s3.Object(bucket_name='bizwarebucket', key=key).get()['Body'].read()
    except ClientError as err:
        logger.error(str(err))


def s3DownloadAllFilesFromFolder(bucket_name='bizwarebucket'):

    try:
        s3 = getS3Resource()

        bucket = s3.Bucket(bucket_name)

        key = 'product/myproject/2021-02-15/'
        objs = list(bucket.objects.filter(Prefix=key))

        for obj in objs:
            # print(obj.key)

            # remove the file name from the object key
            obj_path = os.path.dirname(obj.key)

            # create nested directory structure
            Path(obj_path).mkdir(parents=True, exist_ok=True)

            # save file with full path locally
            bucket.download_file(obj.key, obj.key)

    except botocore.exceptions.ClientError as e:
        if e.response['Error']['Code'] == "404":
            print("The object does not exist.")
        else:
            raise



def s3_download_file(bucket_name='bizwarebucket', key='Sales_Report_Non SAP_22nd_Feb.XLSX'):
    try:
        s3 = getS3Resource()
        s3.Bucket(bucket_name).download_file(key, 'Sales_Report_Non SAP_22nd_Feb.XLSX')
        print("The file is downloaded.")
    except botocore.exceptions.ClientError as e:
        if e.response['Error']['Code'] == "404":
            print("The object does not exist.")
        else:
            raise


def excelFileReader(s3, bucket_name, object_key):
    try:
        # obj = s3.get_object(Bucket=bucket_name, Key=object_key)
        obj = s3.Object(bucket_name='bizwarebucket', key=object_key).get()
        # obj = s3.Object(bucket_name='bizwarebucket', key=file_name).get()['Body'].read()
    except:
        print("That key '" + object_key + "' doesn't exist or unable to connect to s3 for some reason")

    body = obj['Body'].read()
    workbook = load_workbook(io.BytesIO(body), data_only=True)
    workbook.save('com/bizware/data/' + object_key)
    get_sheet_names = workbook.sheetnames
    sheet_index = get_sheet_names.index(get_sheet_names[0])
    workbook.active = sheet_index
    sheet = workbook.active
    print("Sheet Names Available " + str(get_sheet_names))
    print("Sheet Active " + str(sheet))
    max_row = sheet.max_row
    max_col = sheet.max_column
    print("Max Rows " + str(max_row))
    print("Max Columns " + str(max_col))
    # print("===============values_only = FALSE=========")
    # for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=sheet.max_column, values_only=False):
    #     print(row)
    # print("===============values_only = TRUE=========")
    # for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=sheet.max_column, values_only=True):
    #     print(row)
    # print("=========Ignoring the first row=============")
    # for row in sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=sheet.max_column, values_only=True):
    #     print(row)


def s3_download_local(bucket_name, file_name: str, local_path):
    try:
        s3 = getS3Client()
        s3Resource = getS3Resource()
        # file_path = Path.joinpath(local_path, file_name)
        # file_path.parent.mkdir(parents=True, exist_ok=True)
        # return s3.download_file(
        #     bucket_name,
        #     file_name,
        #     str(local_path)
        # )
        return s3Resource.Object(bucket_name='bizwarebucket', key=file_name).get()['Body'].read()
    except ClientError as err:
        logger.error(str(err))


from pathlib import Path


async def get_file_folders(s3_client, bucket_name, prefix=""):
    file_names = []
    folders = []

    default_kwargs = {
        "Bucket": bucket_name,
        "Prefix": prefix
    }
    next_token = ""

    while next_token is not None:
        updated_kwargs = default_kwargs.copy()
        if next_token != "":
            updated_kwargs["ContinuationToken"] = next_token

        response = s3_client.list_objects_v2(**default_kwargs)
        contents = response.get("Contents")

        for result in contents:
            key = result.get("Key")
            if key[-1] == "/":
                folders.append(key)
            else:
                file_names.append(key)

        next_token = response.get("NextContinuationToken")

    return file_names, folders


async def download_files(s3_client, bucket_name, local_path, file_names, folders):
    local_path = Path(local_path)

    for folder in folders:
        folder_path = Path.joinpath(local_path, folder)
        folder_path.mkdir(parents=True, exist_ok=True)

    for file_name in file_names:
        file_path = Path.joinpath(local_path, file_name)
        file_path.parent.mkdir(parents=True, exist_ok=True)
        s3_client.download_file(
            bucket_name,
            file_name,
            str(file_path)
        )


def main():
    client = boto3.client("s3")
    s3Resource = getS3Resource()
    # file_names, folders = get_file_folders(client, "bizwarebucket")
    # download_files(
    #     client,
    #     "bizwarebucket",
    #     "com/bizware/data",
    #     file_names,
    #     folders)

    # s3_download('Sales_Report_Non SAP_22nd_Feb.XLSX')
    # downloadStatus = s3_download_local("bizwarebucket", 'Sales_Report_Non SAP_22nd_Feb.XLSX', 'com/bizware/data')
    # print(downloadStatus)
    s3_download_file()

    # excelFileReader(s3Resource, "bizwarebucket", 'Sales_Report_Non SAP_22nd_Feb.XLSX')


async def create_bucket(bucket_name, region=None):
    """Create an S3 bucket in a specified region

    If a region is not specified, the bucket is created in the S3 default
    region (us-east-1).

    :param bucket_name: Bucket to create
    :param region: String region to create bucket in, e.g., 'us-west-2'
    :return: True if bucket created, else False
    """

    # Create bucket
    try:
        if region is None:
            s3_client = boto3.client('s3')
            s3_client.create_bucket(Bucket=bucket_name)
        else:
            s3_client = boto3.client('s3', region_name=region)
            location = {'LocationConstraint': region}
            s3_client.create_bucket(Bucket=bucket_name,
                                    CreateBucketConfiguration=location)
    except ClientError as e:
        logging.error(e)
        return False
    return True


def upload_file(file_name, bucket, object_name=None):
    """Upload a file to an S3 bucket

    :param file_name: File to upload
    :param bucket: Bucket to upload to
    :param object_name: S3 object name. If not specified then file_name is used
    :return: True if file was uploaded, else False
    """

    # If S3 object_name was not specified, use file_name
    if object_name is None:
        object_name = os.path.basename(file_name)

    # Upload the file
    s3_client = boto3.client('s3')
    try:
        response = s3_client.upload_file(file_name, bucket, object_name)
    except ClientError as e:
        logging.error(e)
        return False
    return True


if __name__ == "__main__":
    main()
